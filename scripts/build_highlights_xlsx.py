from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E1F2"
GREY = "F2F2F2"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "代码亮点存档"
ws.sheet_view.showGridLines = False

widths = [6, 22, 13, 40, 22, 60]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

NCOL = 6

def merge_row(r, value, *, fill=None, font=None, align=None, height=None):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    c = ws.cell(row=r, column=1, value=value)
    if fill:
        c.fill = fill
    if font:
        c.font = font
    c.alignment = align or Alignment(horizontal="left", vertical="center", wrap_text=True)
    if height:
        ws.row_dimensions[r].height = height
    for col in range(1, NCOL + 1):
        ws.cell(row=r, column=col).border = border
    return c

# 标题
merge_row(1, "DPOMBaseMCPServer 代码亮点评比存档",
          fill=PatternFill("solid", fgColor=NAVY),
          font=Font(name=FONT, size=16, bold=True, color=WHITE),
          align=Alignment(horizontal="center", vertical="center"),
          height=34)
merge_row(2, "（优秀代码三项·正式档案）",
          fill=PatternFill("solid", fgColor=NAVY),
          font=Font(name=FONT, size=10, color=WHITE),
          align=Alignment(horizontal="center", vertical="center"),
          height=18)

# 档案信息
info = [
    ("归档项目", "DPOMBaseMCPServer（华为云 AOM/APM/CES 只读监控 MCP Server）"),
    ("评比维度", "软件构造 / 设计模式 / 框架运用"),
    ("入选数量", "3 项（另含荣誉提名 1 项）"),
    ("归档日期", "2026-06-25"),
]
r = 3
for k, v in info:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    kc = ws.cell(row=r, column=1, value=k)
    kc.fill = PatternFill("solid", fgColor=LIGHT)
    kc.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    kc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=NCOL)
    vc = ws.cell(row=r, column=3, value=v)
    vc.font = Font(name=FONT, size=10)
    vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col in range(1, NCOL + 1):
        ws.cell(row=r, column=col).border = border
    ws.row_dimensions[r].height = 20
    r += 1

# 表头
headers = ["序号", "亮点名称", "评比维度", "涉及文件", "核心技术 / 模式", "关键优点"]
hr = r
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=hr, column=i, value=h)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[hr].height = 26
r = hr + 1

rows = [
    ("1",
     "华为云 SDK 容错调用三件套",
     "设计模式 + 框架运用",
     "agentic-common/.../resilience/HuaweiCloudInvocation.java\n"
     "agentic-common/.../sdk/SdkExceptionMapper.java\n"
     "agentic-common/.../resilience/ResilienceConfig.java",
     "Decorator 装饰器模式；Resilience4j 限流/重试；Spring Boot starter 自动装配；ErrorCode 枚举",
     "1. Decorator 模式标准运用：限流→调用→异常映射→重试声明式串成装饰链，各横切关注点独立可组合；装饰顺序保证每次重试都重新过限流闸门，防止重试风暴击穿限流。\n"
     "2. “映射先于重试”：异常在进入重试层前即映射为带 ErrorCode 的 SmartomException，使“是否可重试”收敛到 ErrorCode 枚举这一单一真相源，杜绝多处重复判断与逻辑漂移。\n"
     "3. 优雅复用框架：以 @PostConstruct 改写自动装配的 RetryRegistry 而非新建 Bean，规避循环依赖；退避参数留在 application.yml，配置归配置、逻辑归代码。\n"
     "4. 可观测性与安全内建：成功/失败双路径均记录 api、耗时、错误码、upstreamTraceId；AK/SK 等敏感信息绝不入日志。"),
    ("2",
     "诊断编排服务 DiagnoseTraceService",
     "软件构造 / 并发设计",
     "agentic-monitoring/.../apm/DiagnoseTraceService.java\n"
     "（同类：monitoring/correlate/CorrelateIncidentService.java）",
     "JDK 21 虚拟线程；CompletableFuture fan-out/fan-in；try-with-resources；LinkedHashMap 去重保序",
     "1. 虚拟线程与 CompletableFuture 正确组合：try-with-resources 管理虚拟线程 executor，对同步阻塞 SDK 直接并发，无需自建线程池，安全释放且贴合既定并发模型。\n"
     "2. 两段式 fan-out/fan-in 流水线：先并发拉取拓扑与全部事件，再对筛出的可疑 span 二次并发下钻详情与堆栈/SQL 全文，是具阶段依赖的真实诊断流水线。\n"
     "3. 局部失败不熔断整体：任一阶段失败仅记一条 StageError 并以 null 占位，聚合包同时返回结果与各阶段错误；关联服务更做成 success/failure/skipped 三态。\n"
     "4. 构造细节扎实：以 LinkedHashMap + putIfAbsent 实现可疑 span 去重保序（出错优先、其次最慢）；抽取小工具方法使方法体均控制在规范行数内。"),
    ("3",
     "MCP 工具边界错误翻译层 ToolCallSupport",
     "框架运用 + 健壮性边界",
     "agentic-mcp/.../tool/ToolCallSupport.java",
     "Spring AI MCP Server；高阶函数 Supplier 封装；结构化 ErrorResponse；包私有工具类",
     "1. 深刻理解框架行为：Spring AI MCP 对工具异常仅回传一行 message，Agent 无法获知错误码/可重试性；本类刻意不抛异常而统一转为结构化 ErrorResponse，体现面向“大模型消费者”的 API 设计。\n"
     "2. 进程边界有意例外且记录在案：最外层兜底捕获 RuntimeException 作为出口防线，Javadoc 显式声明为规范例外；完整堆栈仅入服务端日志，回给 Agent 仅含类名与简述，安全与可用性兼顾。\n"
     "3. 高阶函数消除重复模板：以静态 execute(toolName, Supplier) 封装“解析入参→调 service→统一异常→耗时埋点”，包私有 + 私有构造防误用，在数十个工具类中消除重复 try-catch。"),
]

for row in rows:
    for i, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=val)
        c.border = border
        if i == 1:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        elif i in (2, 3):
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.font = Font(name=FONT, size=10, bold=(i == 2))
        else:
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.font = Font(name=FONT, size=9.5)
    fill = PatternFill("solid", fgColor=GREY) if int(row[0]) % 2 == 0 else PatternFill("solid", fgColor=WHITE)
    for col in range(1, NCOL + 1):
        ws.cell(row=r, column=col).fill = fill
    ws.row_dimensions[r].height = 168
    r += 1

# 综合点评
r += 1
merge_row(r, "综合点评",
          fill=PatternFill("solid", fgColor=BLUE),
          font=Font(name=FONT, size=11, bold=True, color=WHITE),
          align=Alignment(horizontal="left", vertical="center"),
          height=24)
r += 1
review = ("三处代码恰好串成一条纵深防御链：HuaweiCloudInvocation 守住上游 SDK 边界，"
          "DiagnoseTraceService 在中间编排层实现局部容错，ToolCallSupport 守住面向 Agent 的出口边界。"
          "每一层都把“失败”翻译成上层可消费的结构化语义，并统一收口到 ErrorCode 枚举。"
          "三者共同体现同一种工程审美：用类型与枚举承载语义、善用框架特性而非蛮力解决问题、"
          "并把“为什么这样写”用中文 Javadoc 沉淀下来。")
merge_row(r, review,
          font=Font(name=FONT, size=10),
          align=Alignment(horizontal="left", vertical="top", wrap_text=True),
          height=86)

# 荣誉提名
r += 1
merge_row(r, "荣誉提名",
          fill=PatternFill("solid", fgColor=LIGHT),
          font=Font(name=FONT, size=10, bold=True, color=NAVY),
          align=Alignment(horizontal="left", vertical="center"),
          height=20)
r += 1
honor = ("契约测试 + 无损 DTO 投影机制（如 AomListMetricsContractTest.java）：以真实 SDK 样本 JSON 反序列化做契约兜底、"
         "反射校验字段，把“SDK 升级悄悄删字段”变为编译期/测试期可见的失败。")
merge_row(r, honor,
          font=Font(name=FONT, size=10),
          align=Alignment(horizontal="left", vertical="top", wrap_text=True),
          height=56)

ws.sheet_view.zoomScale = 110

out = "/Users/huangxinqi/Coding/DPOMBaseMCPServer/代码亮点评比存档.xlsx"
wb.save(out)
print(out)
